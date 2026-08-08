#!/usr/bin/env python3
"""Format botanical names and girth according to Papervale standards."""

import re
from pathlib import Path
from openpyxl import load_workbook

root = Path(__file__).parent.parent
xlsx_file = root / "files" / "availability-list-2026.xlsx"

def format_botanical_name(name):
    """
    Format botanical name: Genus capital, species lowercase, variant quoted with capitals.
    Examples: Abies alba, Acer campestre 'Elsrijk'
    """
    if not name or not isinstance(name, str):
        return name

    name = name.strip()

    # Check if name contains a variant (quoted text)
    variant_match = re.search(r"\s*'([^']+)'", name)
    has_variant = bool(variant_match)

    if has_variant:
        # Extract parts before the variant
        variant_text = variant_match.group(1)
        before_variant = name[:variant_match.start()].strip()
        parts = before_variant.split()

        if len(parts) >= 2:
            # Handle 'x' specially (hybrid indicator)
            genus = parts[0].lower() if parts[0].lower() == 'x' else parts[0].capitalize()
            species = parts[1].lower()
            # Capitalize each word in the variant
            variant_words = variant_text.split()
            variant_capitalized = ' '.join(w.capitalize() for w in variant_words)
            return f"{genus} {species} '{variant_capitalized}'"
        elif len(parts) == 1:
            # Single word with variant (rare)
            genus = parts[0].lower() if parts[0].lower() == 'x' else parts[0].capitalize()
            variant_words = variant_text.split()
            variant_capitalized = ' '.join(w.capitalize() for w in variant_words)
            return f"{genus} '{variant_capitalized}'"
    else:
        # No variant - just format genus and species
        parts = name.split()
        if len(parts) >= 2:
            # Handle 'x' specially (hybrid indicator)
            genus = parts[0].lower() if parts[0].lower() == 'x' else parts[0].capitalize()
            species = parts[1].lower()
            return f"{genus} {species}"
        elif len(parts) == 1:
            # Single word (genus only)
            return parts[0].capitalize()

    return name


def should_remove_value(value):
    """Check if value should be removed (contains 0 in problematic positions)."""
    if not value or value == '':
        return False

    value_str = str(value).strip()
    if not value_str:
        return False

    # Remove if contains patterns like .0, 0., 0', '0
    if re.search(r'\.0|0\.|0\'|\'0', value_str):
        return True

    return False


# Load and update Excel file
print("Loading XLSX file...")
wb = load_workbook(xlsx_file)
ws = wb.active

# Process rows starting from row 3 (row 2 has headers)
updated_names = 0
removed_heights = 0
removed_girths = 0

for row_idx in range(3, ws.max_row + 1):
    # Column B: Botanical Name
    botanical_cell = ws.cell(row=row_idx, column=2)
    # Column E: Height (cm)
    height_cell = ws.cell(row=row_idx, column=5)
    # Column F: Girth (cm)
    girth_cell = ws.cell(row=row_idx, column=6)

    # Format botanical name
    if botanical_cell.value:
        original = botanical_cell.value
        formatted = format_botanical_name(original)
        if formatted != original:
            botanical_cell.value = formatted
            updated_names += 1
            print(f"  Row {row_idx}: {original} → {formatted}")

    # Check height for problematic values
    if should_remove_value(height_cell.value):
        print(f"  Row {row_idx}: Removed height with 0: '{height_cell.value}'")
        height_cell.value = ""
        removed_heights += 1

    # Check girth for problematic values
    if should_remove_value(girth_cell.value):
        print(f"  Row {row_idx}: Removed girth with 0: '{girth_cell.value}'")
        girth_cell.value = ""
        removed_girths += 1

# Save updated Excel
wb.save(xlsx_file)
print(f"\n✓ Excel file updated: {xlsx_file}")
print(f"  Botanical names formatted: {updated_names}")
print(f"  Heights with 0s removed: {removed_heights}")
print(f"  Girths with 0s removed: {removed_girths}")
