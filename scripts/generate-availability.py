#!/usr/bin/env python3
"""Generate PDF availability list from XLSX with logo and repeating headers."""

import json
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageTemplate, Frame, PageBreak
from reportlab.pdfgen import canvas

root = Path(__file__).parent.parent
current_month = datetime.now().strftime('%B').lower()
xlsx_file = root / "files" / f"availability-list-{current_month}-2026.xlsx"
pdf_file = root / "files" / f"availability-list-{current_month}-2026.pdf"
logo_file = root / "assets" / "brand" / "Papervale_LogoMark_Colour_RGB.jpg"

# Load data from XLSX
wb = load_workbook(xlsx_file)
ws = wb.active
rows = []

# Read data from row 5 onwards (rows 1-4 are header section in XLSX)
for row_idx in range(5, ws.max_row + 1):
    sku = ws.cell(row=row_idx, column=1).value
    if sku is None:
        continue
    rows.append({
        'sku': sku,
        'botanical': ws.cell(row=row_idx, column=2).value,
        'common': ws.cell(row=row_idx, column=3).value,
        'pot_size': ws.cell(row=row_idx, column=4).value,
        'height': ws.cell(row=row_idx, column=5).value,
        'girth': ws.cell(row=row_idx, column=6).value,
        'price': ws.cell(row=row_idx, column=7).value,
        'stock': ws.cell(row=row_idx, column=8).value,
        'order': ws.cell(row=row_idx, column=9).value,
    })

date_str = datetime.now().strftime('%-d %B %Y')
total_lines = len(rows)

# Create PDF
doc = SimpleDocTemplate(
    str(pdf_file),
    pagesize=landscape(A4),
    topMargin=30*mm,
    bottomMargin=15*mm,
    leftMargin=10*mm,
    rightMargin=10*mm,
    title="Papervale Trees — Availability List 2026"
)

# Build table data with header (matching Excel column order)
header_row = [
    "SKU", "Botanical Name", "Common Name", "Pot Size",
    "Height (cm)", "Girth (cm)", "Price (inc. vat)", "Stock", "Order"
]

# Table rows
table_data = [header_row]
for row in rows:
    # Format price with £ symbol and 2 decimal places
    if row['price']:
        try:
            price_val = float(str(row['price']).replace('£', ''))
            price_str = f"£{price_val:.2f}"
        except (ValueError, TypeError):
            price_str = str(row['price']) if row['price'] else ""
    else:
        price_str = ""

    table_data.append([
        str(row['sku']) if row['sku'] else "",
        row['botanical'] or "",
        row['common'] or "",
        row['pot_size'] or "",
        str(row['height']) if row['height'] else "",
        str(row['girth']) if row['girth'] else "",
        price_str,
        str(row['stock']) if row['stock'] else "",
        ""  # Order column (blank)
    ])

# Create table with styling
table = Table(table_data, repeatRows=1)
table.setStyle(TableStyle([
    # Header styling
    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#334832')),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
    ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('FONTSIZE', (0, 0), (-1, 0), 8),
    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
    ('TOPPADDING', (0, 0), (-1, 0), 6),

    # Body styling
    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
    ('FONTSIZE', (0, 1), (-1, -1), 8),
    ('TOPPADDING', (0, 1), (-1, -1), 4),
    ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e8e6e4')),

    # Alternating row colors
    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f7f5')]),

    # Right align price column (column 6)
    ('ALIGN', (6, 1), (6, -1), 'RIGHT'),

    # Stock column styling (column 7)
    ('ALIGN', (7, 1), (7, -1), 'CENTER'),
    ('FONTNAME', (7, 1), (7, -1), 'Helvetica-Bold'),

    # Order column styling (column 8) with slightly darker green for visibility
    ('ALIGN', (8, 1), (8, -1), 'CENTER'),
    ('BACKGROUND', (8, 1), (8, -1), colors.HexColor('#B8D4B8')),
]))

# Build story with header
story = []

# Custom header with logo
def header_footer(canvas, doc):
    canvas.saveState()

    # Logo
    if logo_file.exists():
        canvas.drawImage(str(logo_file), 10*mm, landscape(A4)[1] - 25*mm, width=22*mm, height=22*mm)

    # Title and info
    canvas.setFont("Helvetica-Bold", 14)
    canvas.drawString(40*mm, landscape(A4)[1] - 17*mm, "Papervale Trees")

    canvas.setFont("Helvetica", 8)
    canvas.drawString(40*mm, landscape(A4)[1] - 22*mm,
                      f"Spring / Summer 2026 Availability · {total_lines} stock lines · papervaletrees.com")

    # Right side metadata
    canvas.setFont("Helvetica", 8)
    canvas.drawRightString(landscape(A4)[0] - 10*mm, landscape(A4)[1] - 17*mm, f"Generated {date_str}")
    canvas.drawRightString(landscape(A4)[0] - 10*mm, landscape(A4)[1] - 22*mm, "028 3085 0059 · info@papervaletrees.com")

    canvas.restoreState()

# Create a page template with the header
class HeaderFooterPageTemplate(PageTemplate):
    def on_page(self, canvas, doc):
        header_footer(canvas, doc)

doc.build(
    [table],
    onFirstPage=header_footer,
    onLaterPages=header_footer
)

print(f"✓ PDF generated: {pdf_file}")
print(f"  Total stock lines: {total_lines}")
print(f"  File size: {pdf_file.stat().st_size / 1024:.1f} KB")
