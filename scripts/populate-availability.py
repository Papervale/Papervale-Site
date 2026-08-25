#!/usr/bin/env python3
"""
Fetch tree products from Ecwid and populate availability-list-2026.xlsx

IMPORTANT REQUIREMENTS:
- Preserve existing column order: SKU, Botanical Name, Common Name, Pot Size, Height (cm), Girth (cm), Price (GBP), Stock
- Do NOT modify, reformat, or clean tree names — keep exactly as stored in Lightspeed/Ecwid
- Do NOT add or remove columns
- Tree names are the source of truth for botanical nomenclature — preserve all formatting, casing, special characters

Example: "ACACIA melanoxylon / Australian Blackwood" stays exactly as-is (do not split or transform)
"""

import os
import requests
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

root = Path(__file__).parent.parent
env_file = root / ".env"
if env_file.exists():
    with open(env_file) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, value = line.split('=', 1)
                os.environ[key.strip()] = value.strip()

STORE_ID = "73482057"
SECRET_TOKEN = os.getenv('ECWID_SECRET_TOKEN')
if not SECRET_TOKEN:
    print("✗ Error: ECWID_SECRET_TOKEN environment variable not set")
    print("  Add your token to .env file: ECWID_SECRET_TOKEN=your_secret_token")
    exit(1)
BASE_URL = f"https://app.ecwid.com/api/v3/{STORE_ID}"

root = Path(__file__).parent.parent
from datetime import datetime
current_month = datetime.now().strftime('%B').lower()
xlsx_file = root / "files" / f"availability-list-{current_month}-2026.xlsx"

def fetch_products():
    """Fetch all products from Ecwid API."""
    headers = {"Authorization": f"Bearer {SECRET_TOKEN}"}
    products = []
    offset = 0
    limit = 100

    print("Fetching products from Ecwid...")

    while True:
        url = f"{BASE_URL}/products?offset={offset}&limit={limit}"
        response = requests.get(url, headers=headers, timeout=10)

        if response.status_code != 200:
            print(f"✗ API Error: {response.status_code}")
            print(f"Response: {response.text}")
            return None

        data = response.json()
        batch = data.get('items', [])

        if not batch:
            break

        products.extend(batch)
        offset += limit
        print(f"  Fetched {len(products)} products so far...")

    print(f"✓ Total products fetched: {len(products)}")
    return products

def create_availability_xlsx(products):
    """Create XLSX with variant data from Ecwid combinations.

    CRITICAL: Column order and names must match existing file structure.
    Tree names from Ecwid must be preserved exactly — no reformatting or modification.
    NOTE: Gift card and non-tree products are excluded. Each combination variant
          gets its own row. Only non-zero quantity variants are included.
    """
    from datetime import datetime

    wb = Workbook()
    ws = wb.active
    ws.title = "Availability"

    # Add header section with logo and title
    current_month = datetime.now().strftime('%B')
    date_str = datetime.now().strftime('%-d %B %Y')

    # Add logo image
    logo_path = root / "assets" / "brand" / "Papervale_LogoMark_Colour_RGB.jpg"
    if logo_path.exists():
        try:
            img = XLImage(str(logo_path))
            img.width = 40
            img.height = 40
            ws.add_image(img, 'A1')
            ws.row_dimensions[1].height = 45
        except Exception as e:
            print(f"  Warning: Could not insert logo: {e}")

    # Row 1: Title and info (shifted to account for logo)
    ws.merge_cells('C1:I1')
    title_cell = ws.cell(row=1, column=3)
    title_cell.value = "Papervale Trees"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Row 2: Subtitle
    ws.merge_cells('A2:I2')
    subtitle_cell = ws.cell(row=2, column=1)
    subtitle_cell.value = f"Spring / Summer 2026 Availability · Generated {date_str} · papervaletrees.com · 028 3085 0059"
    subtitle_cell.font = Font(size=9)
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 15

    # Add blank row for spacing
    ws.row_dimensions[3].height = 5

    # Table header row (now at row 4)
    headers = ["SKU", "Botanical Name", "Common Name", "Pot Size", "Height (cm)", "Girth (cm)", "Price (inc. vat)", "Stock", "Order"]
    header_fill = PatternFill(start_color="334832", end_color="334832", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = header_border

    # Build data array, filter out gift cards and non-trees, then sort by botanical name
    data_rows = []

    for product in products:
        full_name = product.get('name', '')  # Original from Ecwid

        # Skip gift cards and non-tree products
        if 'gift' in full_name.lower():
            continue

        # Split into botanical and common names
        parts = full_name.split(' / ')
        botanical = parts[0].strip() if parts else full_name
        common = parts[1].strip() if len(parts) > 1 else ""

        # Get combinations (variants) from product
        combinations = product.get('combinations', [])

        if not combinations:
            # If no combinations, skip (we need variant data)
            continue

        # Process each combination/variant
        for combo in combinations:
            quantity = combo.get('quantity', 0)

            # Only include variants with qty > 0
            if quantity <= 0:
                continue

            # Get SKU from combination
            combo_sku = combo.get('sku', '').replace(' - Base', '')

            # Get price from combination
            price = combo.get('defaultDisplayedPrice', 0)

            # Extract pot size, height, and girth from combination options
            pot_size = ""
            height = ""
            girth = ""

            options = combo.get('options', [])
            for opt in options:
                opt_name = opt.get('name', '').lower()
                opt_value = opt.get('value', '')
                if 'pot' in opt_name:
                    pot_size = opt_value
                elif 'height' in opt_name:
                    height = opt_value
                elif 'girth' in opt_name or 'circumference' in opt_name:
                    girth = opt_value

            # Height: trim whitespace only, keep exactly as provided, add "cm" suffix
            formatted_height = ""
            if height:
                h_str = str(height).strip()
                # Add "cm" if not already present
                if h_str and not h_str.lower().endswith('cm'):
                    formatted_height = f"{h_str}cm"
                else:
                    formatted_height = h_str

            # Girth: remove zero-only values (0, .0, 0., 0', 0`, .), etc.), trim whitespace for others
            formatted_girth = ""
            if girth:
                val_str = str(girth).strip()

                # Strip quotes, backticks, spaces, dots to get clean value for checking
                val_clean = val_str.strip().strip("'\"`.").strip().rstrip('.').strip()

                # Check if it's just zero (in any format) or empty/invalid characters
                is_zero = val_clean == '0' or val_clean == '' or not val_clean

                if not is_zero:
                    formatted_girth = val_str

            data_rows.append({
                'sku': combo_sku,
                'botanical': botanical,
                'common': common,
                'pot_size': pot_size,
                'height': formatted_height,
                'girth': formatted_girth,
                'price': price,
                'quantity': quantity,
            })

    # Sort by botanical name
    data_rows.sort(key=lambda x: (x['botanical'] or '').lower())

    # Slightly darker light green for Order column for better visibility
    order_fill = PatternFill(start_color="B8D4B8", end_color="B8D4B8", fill_type="solid")

    # Border for Order column
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Write sorted rows to XLSX (starting at row 5, after header section)
    row_idx = 5
    for row_data in data_rows:
        ws.cell(row=row_idx, column=1).value = row_data['sku']
        ws.cell(row=row_idx, column=2).value = row_data['botanical']
        ws.cell(row=row_idx, column=3).value = row_data['common']
        ws.cell(row=row_idx, column=4).value = row_data['pot_size']
        ws.cell(row=row_idx, column=5).value = row_data['height']
        ws.cell(row=row_idx, column=6).value = row_data['girth']
        ws.cell(row=row_idx, column=7).value = f"£{row_data['price']:.2f}"
        ws.cell(row=row_idx, column=8).value = row_data['quantity']

        # Order column with light green background and visible borders
        order_cell = ws.cell(row=row_idx, column=9)
        order_cell.value = ""
        order_cell.fill = order_fill
        order_cell.border = thin_border

        row_idx += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 12  # SKU
    ws.column_dimensions['B'].width = 25  # Botanical Name
    ws.column_dimensions['C'].width = 25  # Common Name
    ws.column_dimensions['D'].width = 12  # Pot Size
    ws.column_dimensions['E'].width = 12  # Height (cm)
    ws.column_dimensions['F'].width = 12  # Girth (cm)
    ws.column_dimensions['G'].width = 12  # Price (inc. vat)
    ws.column_dimensions['H'].width = 10  # Stock
    ws.column_dimensions['I'].width = 10  # Order

    # Freeze header rows (rows 1-4) so they stay visible when scrolling
    ws.freeze_panes = 'A5'

    # Save
    wb.save(str(xlsx_file))
    print(f"✓ XLSX created: {xlsx_file}")
    print(f"  Total rows: {row_idx - 5}")

    return row_idx - 5

def main():
    """Main flow."""
    print("=" * 60)
    print("Papervale Trees — Populate Availability List")
    print("=" * 60 + "\n")

    # Fetch products
    products = fetch_products()
    if not products:
        print("✗ Failed to fetch products")
        return False

    # Create XLSX
    count = create_availability_xlsx(products)

    print("\n" + "=" * 60)
    print("✓ Next step: Run generate-availability.py to create PDF")
    print("=" * 60)

    return True

if __name__ == "__main__":
    try:
        success = main()
        exit(0 if success else 1)
    except Exception as e:
        print(f"✗ Error: {e}")
        exit(1)
